import os
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, AreaChart, PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select

from app.core.database import get_db
from app.api.deps import get_current_user, get_current_workspace_id
from app.models.user import User
from app.models.dataset import Dataset

router = APIRouter()

def style_range(ws, cell_range, font=None, fill=None, border=None, alignment=None):
    """
    Helper function to style a range of cells in openpyxl.
    """
    for row in ws[cell_range]:
        for cell in row:
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment

@router.get("/download-csv")
async def download_csv(
    dataset_uuid: str,
    current_user: User = Depends(get_current_user),
    workspace_id: int = Depends(get_current_workspace_id),
    db: AsyncSession = Depends(get_db)
):
    """
    Loads the target healed/processed dataset, converts it to CSV format,
    and returns a downloadable file stream.
    """
    result = await db.execute(
        select(Dataset).where(
            Dataset.uuid == dataset_uuid,
            Dataset.workspace_id == workspace_id
        )
    )
    dataset = result.scalars().first()
    if not dataset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Dataset not found or unauthorized access."
        )

    if not os.path.exists(dataset.storage_path):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Processed dataset file not found in storage."
        )

    try:
        if dataset.storage_path.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(dataset.storage_path)
        else:
            df = pd.read_csv(dataset.storage_path)

        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)
        csv_buffer.seek(0)
        
        response = StreamingResponse(
            io.BytesIO(csv_buffer.getvalue().encode('utf-8')),
            media_type="text/csv"
        )
        response.headers["Content-Disposition"] = 'attachment; filename="healed_dataset.csv"'
        return response
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to compile data metrics: {str(e)}"
        )


@router.get("/export-excel")
async def export_excel(
    dataset_uuid: str,
    current_user: User = Depends(get_current_user),
    workspace_id: int = Depends(get_current_workspace_id),
    db: AsyncSession = Depends(get_db)
):
    """
    Generates a fully-styled, interactive Excel Executive Dashboard (.xlsx)
    with pivot-like summaries, dynamic cross-filtering slicers, and a 5-chart visual grid canvas layout.
    """
    result = await db.execute(
        select(Dataset).where(
            Dataset.uuid == dataset_uuid,
            Dataset.workspace_id == workspace_id
        )
    )
    dataset = result.scalars().first()
    if not dataset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Dataset not found or unauthorized access."
        )

    if not os.path.exists(dataset.storage_path):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Processed dataset file not found in storage."
        )

    try:
        # Load healed dataset
        if dataset.storage_path.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(dataset.storage_path)
        else:
            df = pd.read_csv(dataset.storage_path)

        wb = openpyxl.Workbook()
        
        # 1. Executive Dashboard Page Worksheet
        ws = wb.active
        ws.title = "Executive Dashboard"
        ws.views.sheetView[0].showGridLines = True
        ws.freeze_panes = "A6"

        # 2. Operational Data Worksheet
        ws_data = wb.create_sheet(title="Operational Data")
        ws_data.views.sheetView[0].showGridLines = True
        
        # Write dataset records to Operational Data sheet
        headers = list(df.columns)
        ws_data.append(headers)
        for row in df.itertuples(index=False, name=None):
            cleaned_row = ["" if pd.isna(val) else val for val in row]
            ws_data.append(cleaned_row)

        # Autofit Operational Data column widths
        for col in ws_data.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_data.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # 3. Dynamic Column Resolution Mappings
        branch_col = next((c for c in df.columns if 'branch' in c.lower() or 'store' in c.lower() or 'region' in c.lower() or 'state' in c.lower()), None)
        if not branch_col:
            branch_col = next((c for c in df.columns if df[c].dtype == 'object'), df.columns[0])

        date_col = next((c for c in df.columns if 'date' in c.lower() or 'time' in c.lower() or 'month' in c.lower() or 'year' in c.lower()), None)
        if not date_col:
            date_col = next((c for c in df.columns if 'id' in c.lower() or 'num' in c.lower()), df.columns[0])

        sales_col = next((c for c in df.columns if 'revenue' in c.lower() or 'charge' in c.lower() or 'sales' in c.lower() or 'hour' in c.lower() or 'transit' in c.lower() or 'rate' in c.lower() or 'salary' in c.lower() or 'amount' in c.lower()), df.columns[-1])
        profit_col = next((c for c in df.columns if 'profit' in c.lower() or 'income' in c.lower() or 'gross' in c.lower() or 'focus' in c.lower() or 'cancel' in c.lower() or 'streak' in c.lower() or 'age' in c.lower()), df.columns[min(1, len(df.columns)-1)])
        tax_col = next((c for c in df.columns if 'tax' in c.lower() or 'ticket' in c.lower() or 'alarm' in c.lower() or 'fake' in c.lower() or 'frequency' in c.lower() or 'confidence' in c.lower()), df.columns[min(2, len(df.columns)-1)])

        city_col = next((c for c in df.columns if 'city' in c.lower() or 'location' in c.lower() or 'category' in c.lower() or 'type' in c.lower() or 'name' in c.lower()), df.columns[0])
        qty_col = next((c for c in df.columns if 'qty' in c.lower() or 'quantity' in c.lower() or 'count' in c.lower() or 'num' in c.lower() or 'score' in c.lower() or 'hours' in c.lower()), df.columns[min(3, len(df.columns)-1)])

        # Fetch Excel letters matching resolved columns
        branch_letter = get_column_letter(list(df.columns).index(branch_col) + 1)
        date_letter = get_column_letter(list(df.columns).index(date_col) + 1)
        sales_letter = get_column_letter(list(df.columns).index(sales_col) + 1)
        profit_letter = get_column_letter(list(df.columns).index(profit_col) + 1)
        tax_letter = get_column_letter(list(df.columns).index(tax_col) + 1)
        city_letter = get_column_letter(list(df.columns).index(city_col) + 1)
        qty_letter = get_column_letter(list(df.columns).index(qty_col) + 1)

        # Unique lists for summaries
        branches_raw = sorted([str(x) for x in df[branch_col].dropna().unique()])[:10]
        dates_raw = sorted([str(x) for x in df[date_col].dropna().unique()])[:12]
        cities_raw = sorted([str(x) for x in df[city_col].dropna().unique()])[:8]
        categories_raw = sorted([str(x) for x in df[qty_col].dropna().unique()])[:8]
        if not categories_raw:
            categories_raw = ["Category A", "Category B", "Category C"]

        num_rows = len(df) + 1

        # 4. Summary Cache Worksheet (pivot summary mapping table)
        ws_summary = wb.create_sheet(title="Summary Cache")
        ws_summary.views.sheetView[0].showGridLines = True

        # Table 1: Sales by Branch (A1:B11)
        ws_summary["A1"] = "Branch"
        ws_summary["B1"] = "Sales"
        for idx, val in enumerate(branches_raw):
            row_idx = idx + 2
            ws_summary.cell(row=row_idx, column=1, value=val)
            ws_summary.cell(row=row_idx, column=2, value=f"=SUMIFS('Operational Data'!{sales_letter}2:{sales_letter}{num_rows}, 'Operational Data'!{branch_letter}2:{branch_letter}{num_rows}, A{row_idx}, 'Operational Data'!{date_letter}2:{date_letter}{num_rows}, IF('Executive Dashboard'!$D$5=\"ALL\", \"<>\", 'Executive Dashboard'!$D$5))")

        # Table 2: Profit & Tax by Branch (D1:F11)
        ws_summary["D1"] = "Branch"
        ws_summary["E1"] = "Profit"
        ws_summary["F1"] = "Tax"
        for idx, val in enumerate(branches_raw):
            row_idx = idx + 2
            ws_summary.cell(row=row_idx, column=4, value=val)
            ws_summary.cell(row=row_idx, column=5, value=f"=SUMIFS('Operational Data'!{profit_letter}2:{profit_letter}{num_rows}, 'Operational Data'!{branch_letter}2:{branch_letter}{num_rows}, D{row_idx}, 'Operational Data'!{date_letter}2:{date_letter}{num_rows}, IF('Executive Dashboard'!$D$5=\"ALL\", \"<>\", 'Executive Dashboard'!$D$5))")
            ws_summary.cell(row=row_idx, column=6, value=f"=SUMIFS('Operational Data'!{tax_letter}2:{tax_letter}{num_rows}, 'Operational Data'!{branch_letter}2:{branch_letter}{num_rows}, D{row_idx}, 'Operational Data'!{date_letter}2:{date_letter}{num_rows}, IF('Executive Dashboard'!$D$5=\"ALL\", \"<>\", 'Executive Dashboard'!$D$5))")

        # Table 3: Sales Trend by Date (H1:I13)
        ws_summary["H1"] = "Date"
        ws_summary["I1"] = "Sales"
        for idx, val in enumerate(dates_raw):
            row_idx = idx + 2
            ws_summary.cell(row=row_idx, column=8, value=val)
            ws_summary.cell(row=row_idx, column=9, value=f"=SUMIFS('Operational Data'!{sales_letter}2:{sales_letter}{num_rows}, 'Operational Data'!{branch_letter}2:{branch_letter}{num_rows}, IF('Executive Dashboard'!$B$5=\"ALL\", \"<>\", 'Executive Dashboard'!$B$5), 'Operational Data'!{date_letter}2:{date_letter}{num_rows}, H{row_idx})")

        # Table 4: Quantity by City (K1:L9)
        ws_summary["K1"] = "City"
        ws_summary["L1"] = "Quantity"
        for idx, val in enumerate(cities_raw):
            row_idx = idx + 2
            ws_summary.cell(row=row_idx, column=11, value=val)
            ws_summary.cell(row=row_idx, column=12, value=f"=SUMIFS('Operational Data'!{qty_letter}2:{qty_letter}{num_rows}, 'Operational Data'!{branch_letter}2:{branch_letter}{num_rows}, IF('Executive Dashboard'!$B$5=\"ALL\", \"<>\", 'Executive Dashboard'!$B$5), 'Operational Data'!{date_letter}2:{date_letter}{num_rows}, IF('Executive Dashboard'!$D$5=\"ALL\", \"<>\", 'Executive Dashboard'!$D$5), 'Operational Data'!{city_letter}2:{city_letter}{num_rows}, K{row_idx})")

        # Table 5: Operational Analytics (N1:O9)
        ws_summary["N1"] = "Operational Category"
        ws_summary["O1"] = "Avg Value"
        for idx, val in enumerate(categories_raw):
            row_idx = idx + 2
            ws_summary.cell(row=row_idx, column=14, value=val)
            ws_summary.cell(row=row_idx, column=15, value=f"=IFERROR(AVERAGEIFS('Operational Data'!{sales_letter}2:{sales_letter}{num_rows}, 'Operational Data'!{branch_letter}2:{branch_letter}{num_rows}, IF('Executive Dashboard'!$B$5=\"ALL\", \"<>\", 'Executive Dashboard'!$B$5), 'Operational Data'!{date_letter}2:{date_letter}{num_rows}, IF('Executive Dashboard'!$D$5=\"ALL\", \"<>\", 'Executive Dashboard'!$D$5), 'Operational Data'!{qty_letter}2:{qty_letter}{num_rows}, N{row_idx}), 0)")

        # 5. Populate Option Validation Lists to prevent character limits
        branches_options = ["ALL"] + branches_raw
        dates_options = ["ALL"] + dates_raw
        for idx, val in enumerate(branches_options):
            ws_summary.cell(row=idx+1, column=16, value=val) # Column P
        for idx, val in enumerate(dates_options):
            ws_summary.cell(row=idx+1, column=17, value=val) # Column Q

        # Autofit Summary columns
        for col in ws_summary.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # 6. Build Slicer & Cards Layout on Dashboard Page
        font_title = Font(name="Segoe UI", size=15, bold=True, color="1F2937")
        font_slicer_label = Font(name="Segoe UI", size=9, bold=True, color="4B5563")
        font_label = Font(name="Segoe UI", size=9, bold=True, color="374151")
        font_val_blue = Font(name="Segoe UI", size=13, bold=True, color="2563EB")
        font_val_green = Font(name="Segoe UI", size=13, bold=True, color="059669")
        font_val_amber = Font(name="Segoe UI", size=13, bold=True, color="D97706")

        fill_kpi = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_slicer = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        # Title Block
        ws["B1"] = "EXECUTIVE ANALYSIS REAL-TIME BOARD"
        ws["B1"].font = font_title
        ws.row_dimensions[1].height = 28

        # Interactive Slicer Dropdowns (Branch and Date)
        ws["B5"] = "ALL"
        ws["D5"] = "ALL"
        ws["A5"] = "Branch Slicer:"
        ws["A5"].font = font_slicer_label
        ws["A5"].alignment = align_center
        ws["A5"].fill = fill_slicer
        ws["A5"].border = thin_border

        ws["C5"] = "Date Slicer:"
        ws["C5"].font = font_slicer_label
        ws["C5"].alignment = align_center
        ws["C5"].fill = fill_slicer
        ws["C5"].border = thin_border

        # Slicer cells styling
        ws["B5"].font = font_slicer_label
        ws["B5"].alignment = align_center
        ws["B5"].fill = fill_slicer
        ws["B5"].border = thin_border
        
        ws["D5"].font = font_slicer_label
        ws["D5"].alignment = align_center
        ws["D5"].fill = fill_slicer
        ws["D5"].border = thin_border

        # Bind Dropdown options dynamically referencing Summary ranges
        branch_opt_range = f"'Summary Cache'!$P$1:$P${len(branches_options)}"
        date_opt_range = f"'Summary Cache'!$Q$1:$Q${len(dates_options)}"

        dv_branch = DataValidation(type="list", formula1=branch_opt_range, allow_blank=True)
        ws.add_data_validation(dv_branch)
        dv_branch.add(ws["B5"])

        dv_date = DataValidation(type="list", formula1=date_opt_range, allow_blank=True)
        ws.add_data_validation(dv_date)
        dv_date.add(ws["D5"])

        # KPI Summary Cards (B3:C4, D3:E4, F3:G4)
        ws.merge_cells("B3:C3")
        ws.merge_cells("B4:C4")
        ws["B3"] = "TOTAL SALES"
        ws["B4"] = f"=SUMIFS('Operational Data'!{sales_letter}2:{sales_letter}{num_rows}, 'Operational Data'!{branch_letter}2:{branch_letter}{num_rows}, IF($B$5=\"ALL\", \"<>\", $B$5), 'Operational Data'!{date_letter}2:{date_letter}{num_rows}, IF($D$5=\"ALL\", \"<>\", $D$5))"
        ws["B4"].number_format = "$#,##0.00"

        ws.merge_cells("D3:E3")
        ws.merge_cells("D4:E4")
        ws["D3"] = "GROSS PROFIT"
        ws["D4"] = f"=SUMIFS('Operational Data'!{profit_letter}2:{profit_letter}{num_rows}, 'Operational Data'!{branch_letter}2:{branch_letter}{num_rows}, IF($B$5=\"ALL\", \"<>\", $B$5), 'Operational Data'!{date_letter}2:{date_letter}{num_rows}, IF($D$5=\"ALL\", \"<>\", $D$5))"
        ws["D4"].number_format = "$#,##0.00"

        ws.merge_cells("F3:G3")
        ws.merge_cells("F4:G4")
        ws["F3"] = "TAX"
        ws["F4"] = f"=SUMIFS('Operational Data'!{tax_letter}2:{tax_letter}{num_rows}, 'Operational Data'!{branch_letter}2:{branch_letter}{num_rows}, IF($B$5=\"ALL\", \"<>\", $B$5), 'Operational Data'!{date_letter}2:{date_letter}{num_rows}, IF($D$5=\"ALL\", \"<>\", $D$5))"
        ws["F4"].number_format = "$#,##0.00"

        # Apply Card styling
        style_range(ws, "B3:C3", font=font_label, fill=fill_kpi, border=thin_border, alignment=align_center)
        style_range(ws, "B4:C4", font=font_val_blue, fill=fill_kpi, border=thin_border, alignment=align_center)
        style_range(ws, "D3:E3", font=font_label, fill=fill_kpi, border=thin_border, alignment=align_center)
        style_range(ws, "D4:E4", font=font_val_green, fill=fill_kpi, border=thin_border, alignment=align_center)
        style_range(ws, "F3:G3", font=font_label, fill=fill_kpi, border=thin_border, alignment=align_center)
        style_range(ws, "F4:G4", font=font_val_amber, fill=fill_kpi, border=thin_border, alignment=align_center)

        ws.row_dimensions[3].height = 18
        ws.row_dimensions[4].height = 22

        # 7. Render 5-Chart Visual Grid Canvas (width = 14, height = 10)
        chart_w, chart_h = 14, 10

        # Chart 1: Total by Branch Bar Chart (Top Left - B7)
        c1 = BarChart()
        c1.type = "col"
        c1.style = 10
        c1.title = "Total by Branch"
        c1.y_axis.title = "Sales"
        c1.x_axis.title = "Branch"
        c1.width, c1.height = chart_w, chart_h
        data_ref = Reference(ws_summary, min_col=2, min_row=1, max_row=len(branches_raw)+1)
        cats_ref = Reference(ws_summary, min_col=1, min_row=2, max_row=len(branches_raw)+1)
        c1.add_data(data_ref, titles_from_data=True)
        c1.set_categories(cats_ref)
        ws.add_chart(c1, "B7")

        # Chart 2: Profit & Tax Combo Line Chart (Top Right - J7)
        c2 = BarChart()
        c2.type = "col"
        c2.style = 11
        c2.title = "Profit & Tax Summary"
        c2.y_axis.title = "Value"
        c2.x_axis.title = "Branch"
        c2.width, c2.height = chart_w, chart_h
        data_ref2 = Reference(ws_summary, min_col=5, max_col=6, min_row=1, max_row=len(branches_raw)+1)
        cats_ref2 = Reference(ws_summary, min_col=4, min_row=2, max_row=len(branches_raw)+1)
        c2.add_data(data_ref2, titles_from_data=True)
        c2.set_categories(cats_ref2)
        ws.add_chart(c2, "J7")

        # Chart 3: Temporal Sales Trend Area Chart (Middle Left - B23)
        c3 = AreaChart()
        c3.style = 13
        c3.title = "Temporal Sales Trend"
        c3.y_axis.title = "Sales"
        c3.x_axis.title = "Date"
        c3.width, c3.height = chart_w, chart_h
        data_ref3 = Reference(ws_summary, min_col=9, min_row=1, max_row=len(dates_raw)+1)
        cats_ref3 = Reference(ws_summary, min_col=8, min_row=2, max_row=len(dates_raw)+1)
        c3.add_data(data_ref3, titles_from_data=True)
        c3.set_categories(cats_ref3)
        ws.add_chart(c3, "B23")

        # Chart 4: Quantity Distribution Donut Chart (Middle Right - J23)
        c4 = PieChart()
        c4.type = "doughnut"
        c4.style = 10
        c4.title = "Quantity Distribution by City"
        c4.width, c4.height = chart_w, chart_h
        data_ref4 = Reference(ws_summary, min_col=12, min_row=1, max_row=len(cities_raw)+1)
        cats_ref4 = Reference(ws_summary, min_col=11, min_row=2, max_row=len(cities_raw)+1)
        c4.add_data(data_ref4, titles_from_data=True)
        c4.set_categories(cats_ref4)
        ws.add_chart(c4, "J23")

        # Chart 5: 5th Custom Operational Visual Analytics Chart (Bottom Row - B39)
        c5 = LineChart()
        c5.style = 12
        c5.title = "Operational Category Averages"
        c5.y_axis.title = "Avg Value"
        c5.x_axis.title = "Category"
        c5.width, c5.height = chart_w, chart_h
        data_ref5 = Reference(ws_summary, min_col=15, min_row=1, max_row=len(categories_raw)+1)
        cats_ref5 = Reference(ws_summary, min_col=14, min_row=2, max_row=len(categories_raw)+1)
        c5.add_data(data_ref5, titles_from_data=True)
        c5.set_categories(cats_ref5)
        ws.add_chart(c5, "B39")

        # Set specific Column sizes for clean visualization
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 6
        ws.column_dimensions["J"].width = 16

        # Save workbook to memory buffer stream
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        response = StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["Content-Disposition"] = 'attachment; filename="healed_executive_dashboard.xlsx"'
        return response
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate Excel dashboard: {str(e)}"
        )
